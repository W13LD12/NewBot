import sqlite3
import json
import os
import zipfile
from datetime import datetime, date
from openpyxl import Workbook

DB_PATH = "habits.db"

# === ИНИЦИАЛИЗАЦИЯ БАЗЫ ДАННЫХ ===
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()

        # Напоминания по времени (один на пользователя)
        c.execute("""
            CREATE TABLE IF NOT EXISTS reminders (
                user_id INTEGER PRIMARY KEY,
                time TEXT NOT NULL
            )
        """)

        # Привычки (хранятся в виде JSON)
        c.execute("""
            CREATE TABLE IF NOT EXISTS habits (
                user_id INTEGER,
                habit_name TEXT,
                data TEXT,
                PRIMARY KEY (user_id, habit_name)
            )
        """)

        # История выполнения привычек
        c.execute("""
            CREATE TABLE IF NOT EXISTS habit_logs (
                user_id INTEGER,
                habit_name TEXT,
                value REAL,
                timestamp TEXT
            )
        """)

        # Сутевые логи: всё, что вводится по команде /день
        c.execute("""
            CREATE TABLE IF NOT EXISTS daily_logs (
                user_id INTEGER,
                date TEXT,
                data TEXT,
                PRIMARY KEY (user_id, date)
            )
        """)

        # Кастомные поля (определяются пользователем)
        c.execute("""
            CREATE TABLE IF NOT EXISTS custom_fields (
                user_id INTEGER,
                field_name TEXT,
                field_type TEXT,
                PRIMARY KEY (user_id, field_name)
            )
        """)

        # Питание: каждый продукт, по каждому приёму пищи
        c.execute("""
            CREATE TABLE IF NOT EXISTS nutrition_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                date TEXT,
                meal_name TEXT,
                product_name TEXT,
                weight_grams INTEGER,
                calories REAL,
                protein REAL,
                fat REAL,
                carbs REAL,
                salt REAL,
                sugar REAL,
                fiber REAL
            )
        """)

        conn.commit()

# === ДОБАВЛЕНИЕ И СОХРАНЕНИЕ ДАННЫХ ===
def save_habit(user_id, habit_name, data):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT OR REPLACE INTO habits (user_id, habit_name, data)
            VALUES (?, ?, ?)
        """, (user_id, habit_name, json.dumps(data)))
        conn.commit()

def log_habit_value(user_id, habit_name, value):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO habit_logs (user_id, habit_name, value, timestamp)
            VALUES (?, ?, ?, ?)
        """, (user_id, habit_name, value, timestamp))
        conn.commit()

def save_daily_log(user_id, date, data):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT OR REPLACE INTO daily_logs (user_id, date, data)
            VALUES (?, ?, ?)
        """, (user_id, date, json.dumps(data)))
        conn.commit()

def save_custom_field(user_id, field_name, field_type):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT OR REPLACE INTO custom_fields (user_id, field_name, field_type)
            VALUES (?, ?, ?)
        """, (user_id, field_name, field_type))
        conn.commit()

def save_nutrition_entry(user_id, date, meal_name, product_name, grams, calories, protein, fat, carbs, salt, sugar, fiber):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO nutrition_logs (
                user_id, date, meal_name, product_name, weight_grams,
                calories, protein, fat, carbs, salt, sugar, fiber
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user_id, date, meal_name, product_name, grams, calories, protein, fat, carbs, salt, sugar, fiber))
        conn.commit()

# === ПОЛУЧЕНИЕ СУММАРНЫХ НУТРИЕНТОВ ===
def get_nutrition_summary(user_id: int):
    today = date.today().isoformat()
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                COALESCE(SUM(calories), 0),
                COALESCE(SUM(protein), 0),
                COALESCE(SUM(fat), 0),
                COALESCE(SUM(carbs), 0),
                COALESCE(SUM(salt), 0),
                COALESCE(SUM(sugar), 0),
                COALESCE(SUM(fiber), 0)
            FROM nutrition_logs
            WHERE user_id = ? AND date = ?
        """, (user_id, today))
        row = cursor.fetchone()
        return {
            'calories': row[0],
            'protein': row[1],
            'fat': row[2],
            'carbs': row[3],
            'salt': row[4],
            'sugar': row[5],
            'fiber': row[6]
        }

# === ЭКСПОРТ В EXCEL ===
def export_to_excel(user_id, output_path):
    wb = Workbook()
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()

        # Привычки
        c.execute("SELECT habit_name, value, timestamp FROM habit_logs WHERE user_id = ?", (user_id,))
        habits_data = c.fetchall()
        ws_habits = wb.active
        ws_habits.title = "Привычки"
        ws_habits.append(["Название", "Значение", "Время"])
        for row in habits_data:
            ws_habits.append(row)

        # Логи дня
        c.execute("SELECT date, data FROM daily_logs WHERE user_id = ?", (user_id,))
        logs_data = c.fetchall()
        ws_logs = wb.create_sheet("Логи дня")
        ws_logs.append(["Дата", "Данные"])
        for row in logs_data:
            ws_logs.append([row[0], json.loads(row[1])])

        # Питание
        c.execute("""
            SELECT date, meal_name, product_name, weight_grams, calories, protein, fat, carbs, salt, sugar, fiber
            FROM nutrition_logs WHERE user_id = ?
        """, (user_id,))
        food_data = c.fetchall()
        ws_nutrition = wb.create_sheet("Питание")
        ws_nutrition.append([
            "Дата", "Приём пищи", "Продукт", "Граммы", "Ккал",
            "Белки", "Жиры", "Углеводы", "Соль", "Сахар", "Клетчатка"
        ])
        for row in food_data:
            ws_nutrition.append(row)

    wb.save(output_path)

# === РЕЗЕРВНАЯ КОПИЯ ===
def backup_database(zip_name="backup.zip"):
    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(DB_PATH)

# === ЗАГРУЗКА ПОЛЕЙ И ДАННЫХ ===
def get_custom_fields(user_id):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT field_name, field_type FROM custom_fields WHERE user_id = ?", (user_id,))
        return c.fetchall()

def get_daily_log(user_id, date):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT data FROM daily_logs WHERE user_id = ? AND date = ?", (user_id, date))
        row = c.fetchone()
        return json.loads(row[0]) if row else {}

def get_latest_water_log(user_id):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT data FROM daily_logs
            WHERE user_id = ?
            ORDER BY date DESC LIMIT 1
        """, (user_id,))
        row = c.fetchone()
        if row:
            data = json.loads(row[0])
            return data.get("вода")
        return None